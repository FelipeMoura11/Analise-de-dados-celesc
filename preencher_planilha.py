import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Arquivos
ARQUIVO_EXISTENTE = r'C:\Users\User2025\Desktop\CASE\Excel\Case_Nome_Sobrenome.xlsx'
PASTA_SAIDA = r'C:\Users\User2025\Desktop\CASE\output'
ARQUIVO_SAIDA = os.path.join(PASTA_SAIDA, 'Case_Nome_Sobrenome_Preenchido.xlsx')
ARQUIVO_DADOS = os.path.join(PASTA_SAIDA, 'receita_liquida.csv')

def periodo_para_chave(periodo):
    # Exemplo: "1T25" -> trimestre=1, ano=25
    try:
        trimestre = int(periodo[0])
        ano = int(periodo[2:])  # pega o ano depois do 'T'
        # Chave para ordenar decrescente ano e crescente trimestre: 
        # ano * 10 - trimestre (multiplica para não misturar 2 dígitos de ano)
        # Para ano decrescente, usamos valor negativo do ano
        return (ano, trimestre)
    except Exception as e:
        # Caso dê problema, joga pra última posição
        return (9999, 99)

def preencher_planilha_receita(arquivo_dados, arquivo_existente, arquivo_saida):
    df = pd.read_csv(arquivo_dados)
    df = df[['Conta', 'Valor', 'Período']]
    
    # Ordena usando a chave customizada
    df['chave_ordem'] = df['Período'].apply(periodo_para_chave)
    df = df.sort_values(by='chave_ordem').drop(columns=['chave_ordem'])

    wb = load_workbook(arquivo_existente)
    ws = wb['Database']

    start_row = 4
    start_col = 2  # Coluna B

    print(f"➡️ Preenchendo planilha com {len(df)} registros...")

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    if not os.path.exists(PASTA_SAIDA):
        os.makedirs(PASTA_SAIDA)

    wb.save(arquivo_saida)
    print(f"✅ Planilha preenchida e salva como: {arquivo_saida}")

if __name__ == '__main__':
    preencher_planilha_receita(ARQUIVO_DADOS, ARQUIVO_EXISTENTE, ARQUIVO_SAIDA)
